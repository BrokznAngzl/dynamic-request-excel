import json
import re
from datetime import datetime

import openpyxl
import pandas as pd
import requests
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

url = "{url}"
parameter_center_path = '../resource/parameter_center.xlsx'
new_profile_excel_path = '../resource/new_profile.xlsx.xlsx'
request_path = "../resource/request.txt"

token = "{token}"
headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {token}"
}


def get_account_info(response):
    data = response.json()
    return data['accountNo'], data['activeDate']


def get_sys_date():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def write_response_to_excel(account_no, active_date, row):
    try:
        print(f"open file {new_profile_excel_path}")
        wb = openpyxl.load_workbook(new_profile_excel_path)
        sheet = wb["new_profile"]
        sheet.cell(row=row, column=1).value = str(account_no)
        sheet.cell(row=row, column=2).value = str(active_date)
        wb.save(new_profile_excel_path)
        print(f"Saved {account_no} to Excel at row {row}")
    except Exception as e:
        print(f"Failed to write to Excel: {str(e)}")


def get_value_from_excel(row, column):
    # too many open excel
    # need more implement
    df = pd.read_excel(parameter_center_path, sheet_name='request info', dtype=str)
    value = df.at[row, column]
    return "" if pd.isna(value) or str(value).strip() == "" else str(value).strip()


def split_keys(key_string):
    parts = []
    for part in key_string.split('.'):
        matches = re.findall(r'([^\[\]]+)|\[(\d+)\]', part)
        for name, idx in matches:
            if name:
                parts.append(name)
            if idx:
                parts.append(int(idx))
    return parts


def set_nested_value(d, key_string, value):
    # too many condition
    # need more implement
    keys = split_keys(key_string)
    for i, key in enumerate(keys[:-1]):
        next_key = keys[i + 1]

        if isinstance(key, int):
            if not isinstance(d, list):
                print(f"Error: Expected list but got {type(d).__name__} at key {key}")
                raise TypeError(f"Expected list at this level but got {type(d)}")
            while len(d) <= key:
                d.append({} if not isinstance(next_key, int) else [])
            d = d[key]
        else:
            if not isinstance(d, dict):
                print(f"Error: Expected dict but got {type(d).__name__} at key {key}")
                raise TypeError(f"Expected dict at this level but got {type(d)}")
            if key not in d:
                d[key] = [] if isinstance(next_key, int) else {}
            d = d[key]

    last_key = keys[-1]
    if isinstance(last_key, int):
        if not isinstance(d, list):
            print(f"Error: Expected list at last level but got {type(d).__name__}")
            raise TypeError(f"Expected list at last level but got {type(d)}")
        while len(d) <= last_key:
            d.append(None)
        d[last_key] = value
    else:
        if not isinstance(d, dict):
            # print(f"Error: Expected dict at last level but got {type(d).__name__}")
            raise TypeError(f"Expected dict at last level but got {type(d)}")
        d[last_key] = value


def get_payload(record):
    with open(request_path, 'r', encoding='utf-8') as file:
        body = json.load(file)

    df = pd.read_excel(parameter_center_path, sheet_name='request info', dtype=str)
    columns = df.columns

    for col in columns:
        value = get_value_from_excel(record, col)
        set_nested_value(body, col, value)

    preOrderDate = get_value_from_excel(record, 'preOrderDate')
    if not preOrderDate:
        preOrderDate = get_sys_date()

    body['preOrderDate'] = preOrderDate

    return body


def create_account(record):
    try:
        print(f"prepare properties for record: {record}")
        payload = get_payload(record)
        # print(f"payload: {payload}")

        print("sending request...")
        response = requests.post(url, json=payload, headers=headers, verify=False)
        print(f"Status: {response.status_code} - {response.text}")
        account_no, active_date = get_account_info(response)

        excel_row = record + 2
        write_response_to_excel(account_no, active_date, excel_row)

    except Exception as e:
        print(f"Error: {str(e)}")


if __name__ == "__main__":
    profile_records = 10
    for i in range(profile_records):
        create_account(i)
